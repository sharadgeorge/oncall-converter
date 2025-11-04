"""
Cardiology Department Configuration
====================================
Configuration for converting Cardiology Excel schedules to import format.

This file contains all Cardiology-specific logic for:
- File validation
- Data extraction (TO BE COMPLETED)
- Schedule processing

Status: Structure defined, extraction logic needs implementation
To use: Place in department_configs/ folder or same directory as main script.

Note: Utilities (DepartmentConfig, col_letter_to_index, etc.) are injected
by the main script during dynamic loading. No need to import them.
"""

import openpyxl
import calendar
import re
from datetime import datetime, timedelta
from pathlib import Path

# Utilities are injected by main script:
# - DepartmentConfig
# - col_letter_to_index
# - is_weekday
# - extract_month_year_from_filename
# - create_schedule_entry
# - DebugTracker

class CardiologyConfig(DepartmentConfig):
    """Configuration for Cardiology department schedules"""
    
    def get_department_name(self):
        """Return department name"""
        return "Cardiology"
    
    def get_file_requirements(self):
        """Return list of required file descriptions"""
        return [
            "Rotation Schedule file (.xlsx) - for Teams 94 and 123",
            "Team 8 Schedule file (.xlsx) - for Cardiovascular team",
            #"(Optional) Team 8 Hospital file (.xlsx) - for additional assignments"
        ]
    
    def get_team_abbr_map(self):
        """Return team name abbreviations for display"""
        return {
            'Cardiovascular': 'CardioVasc',
            'Interventional Cardiologist': 'Int_Cardio',
            'Cardiology': 'Cardiology'
        }
    
    def validate_and_configure(self, files, month, year):
        """
        Validate Cardiology files and configure settings
        
        Args:
            files: List of Path objects [rotation_file, team8_file, optional_hospital_file]
            month: Initial month (may be None)
            year: Initial year (may be None)
            
        Returns:
            tuple: (workbooks_dict, config_dict)
        """
        if len(files) < 2:
            raise ValueError("Cardiology requires at least 2 files")
        
        rotation_file = files[0]
        team8_file = files[1]
        #hospital_file = files[2] if len(files) > 2 else None
        
        print("\n" + "=" * 60)
        print("Validating Cardiology Schedule Files")
        print("=" * 60)
        
        # Load Rotation file (Teams 94 and 123)
        print(f"\n  Loading Rotation Schedule: {rotation_file.name}")
        wb_rotation = openpyxl.load_workbook(rotation_file, data_only=True)
        print(f"  ✓ Loaded with {len(wb_rotation.sheetnames)} sheets")
        
        # Load Team 8 file
        print(f"\n  Loading Team 8 Schedule: {team8_file.name}")
        wb_team8 = openpyxl.load_workbook(team8_file, data_only=True)
        print(f"  ✓ Loaded with {len(wb_team8.sheetnames)} sheets")
        
        # Load optional hospital file
        #wb_hospital = None
        #if hospital_file:
        #    print(f"\n  Loading Team 8 Hospital file: {hospital_file.name}")
        #    wb_hospital = openpyxl.load_workbook(hospital_file, data_only=True)
        #    print(f"  ✓ Loaded with {len(wb_hospital.sheetnames)} sheets")
        
        # Use month/year from Streamlit parameters if provided
        if month is not None and year is not None:
            # Parameters already set by Streamlit - use them
            print(f"\n  ✓ Processing: {calendar.month_name[month]} {year}")
        else:
            # Fallback: Detect from filename (for command-line use)
            detected_month, detected_year = extract_month_year_from_filename(rotation_file.stem)
            if detected_month:
                print(f"\n  Detected month: {calendar.month_name[detected_month]} {detected_year}")
                month = detected_month
                year = detected_year
                print(f"  ✓ Processing: {calendar.month_name[month]} {year}")
            else:
                # Last fallback: use current date
                month = datetime.now().month
                year = datetime.now().year
                print(f"\n  Using current date: {calendar.month_name[month]} {year}")

        
        # Row configuration - use defaults for Streamlit (no prompts)
        print("\nRow Configuration:")
        print("-" * 60)
        
        # Team 123 Consultant rows
        consultant_rows = (6, 11)
        print(f"✓ Using default rows {consultant_rows[0]}-{consultant_rows[1]} for Team 123 Consultants")
        
        # Team 123 Staff/Fellows rows
        staff_rows = (6, 13)
        print(f"✓ Using default rows {staff_rows[0]}-{staff_rows[1]} for Team 123 Staff/Fellows")
        
        # Team 94 row
        team94_row = 29
        print(f"✓ Using default row {team94_row} for Team 94")
        
        # Team 8 rows
        team8_rows = (12, 16)
        print(f"✓ Using default rows {team8_rows[0]}-{team8_rows[1]} for Team 8")
        
        print()
        
        # Team configurations for Cardiology
        teams = {
            'Cardiovascular': {
                'team_id': '8',
                'data_rows': team8_rows,
                'first_col': 'C',
                'last_col': 'AG'
            },
            'Interventional Cardiologist': {
                'team_id': '94',
                'data_row': team94_row,
                'first_col': 'D',
                'last_col': 'AH'
            },
            'Cardiology': {
                'team_id': '123',
                'consultant_rows': consultant_rows,
                'staff_rows': staff_rows,
                'attending_name_col': 'B',  # Column B for consultant names
                'sp_name_col': 'B',  # Column B for staff/fellow names
                'first_col': 'D',
                'last_col': 'AH'
            }
        }
        
        workbooks = {
            'rotation': wb_rotation,
            'team8': wb_team8,
            #'hospital': wb_hospital
        }
        
        config = {
            'teams': teams,
            'month': month,
            'year': year
        }
        
        return workbooks, config
    
    def extract_schedule_data(self, workbooks, config, month, year):
        """
        Extract Cardiology schedule data from workbooks
        
        Args:
            workbooks: Dict containing worksheet objects
            config: Dict containing configuration
            month: Month number (1-12)
            year: Year (e.g., 2025)
            
        Returns:
            list: List of schedule entry dictionaries
        """
        teams = config['teams']
        
        # Extract data from each source
        print("\nReading Cardiovascular (Team 8) assignments...")
        cardiovascular_data = self._read_cardiovascular_data(workbooks['team8'], teams['Cardiovascular'], month, year)
        cardio_days = len([d for d in cardiovascular_data if cardiovascular_data[d]])
        print(f"  ✓ Found assignments on {cardio_days} days")
        
        print("\nReading Interventional Cardiologist (Team 94) assignments...")
        interventional_data = self._read_interventional_data(workbooks['rotation'], teams['Interventional Cardiologist'], month, year)
        intv_days = len(interventional_data)
        print(f"  ✓ Found assignments on {intv_days} days")
        
        print("\nReading Cardiology (Team 123) assignments...")
        cardiology_data = self._read_cardiology_data(workbooks['rotation'], teams['Cardiology'], month, year)
        cardiology_days = len([d for d in cardiology_data if cardiology_data[d]])
        print(f"  ✓ Found assignments on {cardiology_days} days")
        
        # Debug counts
        consultant_count = sum(len(cardiology_data[d]['consultants']) for d in cardiology_data if cardiology_data[d])
        staff_count = sum(len(cardiology_data[d]['staff']) for d in cardiology_data if cardiology_data[d])
        print(f"  DEBUG: Consultants found: {consultant_count}, Staff found: {staff_count}")
        
        # Create output data
        print("Generating output data...")
        output_data = self._create_output_data(cardiovascular_data, interventional_data, cardiology_data, year, month)
        
        print(f"✓ Generated {len(output_data)} schedule entries")
        
        return output_data
    
    def _read_cardiovascular_data(self, wb, config, month_num, year):
        """Read on-call data from Cardiovascular file - rows with X/XA/XP markers"""
        # Find the appropriate sheet
        ws = None
        for sheet_name in wb.sheetnames:
            if 'on' in sheet_name.lower() and 'call' in sheet_name.lower():
                ws = wb[sheet_name]
                break
        
        if not ws:
            ws = wb.active
        
        # Get column and row ranges
        first_col_idx = col_letter_to_index(config['first_col'])
        last_col_idx = col_letter_to_index(config['last_col'])
        row_start, row_end = config['data_rows']
        
        # Marker to role mapping
        MARKER_TO_ROLES = {
            'X': ['84', '2001'],    # Both Echo Tech Adult and Echo Tech Ped
            'XA': ['84'],           # Echo Tech Adult only
            'XP': ['2001'],         # Echo Tech Ped only
        }
        
        # Dictionary to store assignments: {day: [(username, [roles])]}
        assignments = {}
        days_in_month = calendar.monthrange(year, month_num)[1]
        
        # Read data for each day
        for day in range(1, days_in_month + 1):
            col_idx = first_col_idx + day - 1
            if col_idx > last_col_idx:
                break
            
            assignments[day] = []
            
            # Check each employee row
            for row in range(row_start, row_end + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                
                if cell_value:
                    cell_str = str(cell_value).strip().upper()
                    
                    # Check if this is a valid marker
                    if cell_str in MARKER_TO_ROLES:
                        # Get employee name/initials from column B (index 2)
                        emp_cell = ws.cell(row=row, column=2).value
                        if emp_cell:
                            username = self._find_username_by_identifier(emp_cell)
                            # Store even if username is None (unknown employee)
                            roles = MARKER_TO_ROLES[cell_str]
                            assignments[day].append((username, roles))
                            # NOTE: Expected entries are tracked later during output creation
        
        return assignments
    
    def _read_interventional_data(self, wb, config, month_num, year):
        """Read on-call data from Interventional Cardiologist file"""
        month_name = calendar.month_name[month_num]
        month_abbr = calendar.month_abbr[month_num]
        
        # Find the appropriate sheet
        ws = None
        
        # Try exact match with "{MonthAbbr} Attending"
        attending_pattern = f"{month_abbr} Attending"
        for sheet_name in wb.sheetnames:
            if sheet_name == attending_pattern:
                ws = wb[sheet_name]
                print(f"  ✓ Using sheet for Team 94: '{sheet_name}'")
                break
        
        # If not found, try case-insensitive fuzzy match
        if not ws:
            for sheet_name in wb.sheetnames:
                sheet_lower = sheet_name.lower()
                if month_name.lower() in sheet_lower or month_abbr.lower() in sheet_lower:
                    ws = wb[sheet_name]
                    print(f"  ✓ Using sheet for Team 94: '{sheet_name}'")
                    break
        
        # If still not found, prompt user
        if not ws:
            print(f"\n  ⚠ ERROR: Could not find sheet for Team 94 for {month_name} {year}")
            print(f"  Expected: Sheet containing month name")
            print(f"\n  Available sheets in workbook:")
            for i, sheet_name in enumerate(wb.sheetnames, 1):
                print(f"    {i}. {sheet_name}")
            
            sheet_input = input(f"\n  Enter exact sheet name for Team 94: ").strip()
            if sheet_input and sheet_input in wb.sheetnames:
                ws = wb[sheet_input]
                print(f"  ✓ Using sheet: '{sheet_input}'")
            else:
                print(f"  ⚠ WARNING: Cannot find valid sheet for Team 94, skipping...")
                return {}
        
        # Get column range and row
        first_col_idx = col_letter_to_index(config['first_col'])
        last_col_idx = col_letter_to_index(config['last_col'])
        row = config['data_row']
        
        # Dictionary to store assignments: {day: username}
        assignments = {}
        days_in_month = calendar.monthrange(year, month_num)[1]
        
        # Read data for each day
        for day in range(1, days_in_month + 1):
            col_idx = first_col_idx + day - 1
            if col_idx > last_col_idx:
                break
            
            cell_value = ws.cell(row=row, column=col_idx).value
            
            if cell_value:
                username = self._find_username_by_identifier(cell_value)
                # Store even if username is None (unknown employee)
                assignments[day] = username
                # NOTE: Expected entries are tracked later during output creation
        
        return assignments
    
    def _read_cardiology_data(self, wb, config, month_num, year):
        """Read on-call data from Cardiology (Team 123) file - two worksheets"""
        month_name = calendar.month_name[month_num]
        month_abbr = calendar.month_abbr[month_num]
        
        # Find Attending sheet
        attending_sheet = None
        attending_pattern = f"{month_abbr} Attending"
        for sheet_name in wb.sheetnames:
            if sheet_name == attending_pattern:
                attending_sheet = wb[sheet_name]
                print(f"  ✓ Found Attending sheet: '{sheet_name}'")
                break
        
        # If not found, try fuzzy match
        if not attending_sheet:
            for sheet_name in wb.sheetnames:
                sheet_lower = sheet_name.lower()
                if month_abbr.lower() in sheet_lower and 'attending' in sheet_lower:
                    attending_sheet = wb[sheet_name]
                    print(f"  ✓ Found Attending sheet: '{sheet_name}'")
                    break
                elif month_name.lower() in sheet_lower and 'attending' in sheet_lower:
                    attending_sheet = wb[sheet_name]
                    print(f"  ✓ Found Attending sheet: '{sheet_name}'")
                    break
        
        # Find SP sheet
        sp_sheet = None
        sp_pattern = f"{month_abbr} {year} SP"
        for sheet_name in wb.sheetnames:
            if sheet_name == sp_pattern:
                sp_sheet = wb[sheet_name]
                print(f"  ✓ Found SP sheet: '{sheet_name}'")
                break
        
        # If not found, try fuzzy match
        if not sp_sheet:
            for sheet_name in wb.sheetnames:
                sheet_lower = sheet_name.lower()
                if (month_abbr.lower() in sheet_lower or month_name.lower() in sheet_lower) and 'sp' in sheet_lower:
                    sp_sheet = wb[sheet_name]
                    print(f"  ✓ Found SP sheet: '{sheet_name}'")
                    break
        
        # Prompt if sheets not found
        if not attending_sheet:
            print(f"\n  ⚠ ERROR: Could not find Attending sheet for {month_name} {year}")
            print(f"  Expected pattern: '{month_abbr} Attending'")
            print(f"\n  Available sheets:")
            for i, sheet_name in enumerate(wb.sheetnames, 1):
                print(f"    {i}. {sheet_name}")
            
            sheet_input = input(f"\n  Enter exact Attending sheet name: ").strip()
            if sheet_input and sheet_input in wb.sheetnames:
                attending_sheet = wb[sheet_input]
                print(f"  ✓ Using sheet: '{sheet_input}'")
            else:
                raise ValueError(f"Cannot proceed without valid Attending sheet for Team 123")
        
        if not sp_sheet:
            print(f"\n  ⚠ ERROR: Could not find SP sheet for {month_name} {year}")
            print(f"  Expected pattern: '{month_abbr} {year} SP'")
            print(f"\n  Available sheets:")
            for i, sheet_name in enumerate(wb.sheetnames, 1):
                print(f"    {i}. {sheet_name}")
            
            sheet_input = input(f"\n  Enter exact SP sheet name: ").strip()
            if sheet_input and sheet_input in wb.sheetnames:
                sp_sheet = wb[sheet_input]
                print(f"  ✓ Using sheet: '{sheet_input}'")
            else:
                raise ValueError(f"Cannot proceed without valid SP sheet for Team 123")
        
        # Get column range
        first_col_idx = col_letter_to_index(config['first_col'])
        last_col_idx = col_letter_to_index(config['last_col'])
        days_in_month = calendar.monthrange(year, month_num)[1]
        
        # Dictionary to store assignments per day
        assignments = {}
        
        # Read Attending Consultants
        attending_name_col = col_letter_to_index(config['attending_name_col'])
        row_start, row_end = config['consultant_rows']
        
        print(f"  DEBUG: Reading Attending consultants from rows {row_start}-{row_end}")
        consultant_entries_found = 0
        
        for day in range(1, days_in_month + 1):
            col_idx = first_col_idx + day - 1
            if col_idx > last_col_idx:
                break
            
            if day not in assignments:
                assignments[day] = {'consultants': [], 'staff': []}
            
            # Read consultants for this day
            for row in range(row_start, row_end + 1):
                cell_value = attending_sheet.cell(row=row, column=col_idx).value
                
                if cell_value:
                    # Get consultant name from column B
                    emp_name = attending_sheet.cell(row=row, column=attending_name_col).value
                    if emp_name:
                        username = self._find_username_by_identifier(emp_name)
                        # Parse markers from cell (consultant parsing - strict)
                        markers = self._parse_day_abbreviations(cell_value, is_staff=False)
                        
                        # Add to assignments even if username is None (unknown employee)
                        # This allows us to track expected entries for unknowns
                        if markers:
                            assignments[day]['consultants'].append((username, markers))
                            if username:  # Only count as found if username exists
                                consultant_entries_found += 1
                            if day <= 3:  # Debug first 3 days
                                print(f"    Day {day}, Row {row}: {emp_name} -> {username}, cell: '{cell_value}', markers: {markers}")
                            
                            # NOTE: Expected entries are tracked later during output creation
                            # when we actually decide which consultant to use
        
        print(f"  DEBUG: Total consultant entries found: {consultant_entries_found}")
        
        # Read Staff/Fellows
        sp_name_col = col_letter_to_index(config['sp_name_col'])
        row_start, row_end = config['staff_rows']
        
        for day in range(1, days_in_month + 1):
            col_idx = first_col_idx + day - 1
            if col_idx > last_col_idx:
                break
            
            if day not in assignments:
                assignments[day] = {'consultants': [], 'staff': []}
            
            # Read staff/fellows for this day
            for row in range(row_start, row_end + 1):
                cell_value = sp_sheet.cell(row=row, column=col_idx).value
                
                if cell_value:
                    markers = self._parse_day_abbreviations(cell_value, is_staff=True)
                    
                    if markers:
                        # Get staff/fellow name from column B
                        emp_name = sp_sheet.cell(row=row, column=sp_name_col).value
                        if emp_name:
                            username = self._find_username_by_identifier(emp_name)
                            # Add to assignments even if username is None (unknown employee)
                            assignments[day]['staff'].append((username, markers))
                            
                            # NOTE: Expected entries are tracked later during output creation
                            # when we actually decide which staff to assign to each shift
        
        return assignments
    
    def _parse_day_abbreviations(self, cell_value, is_staff=False):
        """Parse abbreviations in cells and return list of markers found"""
        if not cell_value:
            return []
        
        cell_str = str(cell_value).strip().upper()
        cell_str = cell_str.replace('\n', ' ')
        
        markers = []
        
        # Check for combined markers first
        if is_staff and ('2C/E' in cell_str or '2CE' in cell_str):
            markers.extend(['D', 'E'])
            return markers
        
        # Split by spaces
        parts = cell_str.split()
        
        for part in parts:
            part = part.strip()
            if not part:
                continue
            
            # CONSULTANT markers - only D, LD, DL, D/A, X
            if part == 'D':
                markers.append('D')
            elif part in ['LD', 'DL', 'D/A']:
                markers.append(part)
            elif part == 'X':  # Weekend marker
                markers.append('X')
            
            # STAFF-specific markers (only when is_staff=True)
            elif is_staff:
                if part == '2C':
                    markers.append('2C')
                elif part == 'N':
                    markers.append('N')
                elif part in ['E', 'CE', '2BE']:
                    markers.append('E')
        
        return markers
    
    def _create_output_data(self, cardiovascular_data, interventional_data, cardiology_data, year, month):
        """Create output data structure"""
        output_rows = []
        days_in_month = calendar.monthrange(year, month)[1]
        
        for day in range(1, days_in_month + 1):
            current_date = datetime(year, month, day)
            next_date = current_date + timedelta(days=1)
            
            is_weekday_flag = is_weekday(current_date)
            
            # Process Cardiology team (123) - FIRST in order
            if day in cardiology_data:
                day_data = cardiology_data[day]
                
                if is_weekday_flag:
                    # WEEKDAYS: 4 entries
                    # 1. Consultant 2nd Day Call (7:00-16:00)
                    # 2. Staff/Fellow 1st Day Call (7:00-16:00)
                    # 3. Staff/Fellow Evening Call (16:00-19:00)
                    # 4. Staff/Fellow Night Call (19:00-07:00)
                    
                    # Find consultant
                    consultant_username = None
                    for username, markers in day_data['consultants']:
                        if username and 'D' in markers:  # Skip None values in selection
                            consultant_username = username
                            break
                    
                    if not consultant_username:
                        for username, markers in day_data['consultants']:
                            if username and any(m in markers for m in ['LD', 'DL', 'D/A']):
                                consultant_username = username
                                break
                    
                    if not consultant_username:
                        # Use first consultant that has a valid username
                        for username, markers in day_data['consultants']:
                            if username:
                                consultant_username = username
                                break
                    
                    # Track expected entry if we have ANY consultant data (even if unknown)
                    if day_data['consultants']:
                        self.debug_tracker.add_expected_entry('Cardiology', day, '700', '1600', 'Team123', None)
                        
                        # Only create entry if we have a valid username
                        if consultant_username:
                            entry = create_schedule_entry(consultant_username, '123', current_date, '700', next_date, '1600', '3042457')
                            entry['NOTES'] = '2nd Day Call'
                            output_rows.append(entry)
                            self.debug_tracker.mark_entry_generated('Cardiology', day, '700', '1600')
                    day_call_username = None
                    evening_call_username = None
                    night_call_username = None
                    has_day_assignment = False
                    has_evening_assignment = False
                    has_night_assignment = False
                    
                    # Priority 1: Look for 'D' marker first (skip None values)
                    for username, markers in day_data['staff']:
                        if username and 'D' in markers and not day_call_username:
                            day_call_username = username
                            has_day_assignment = True
                        if username and 'E' in markers and not evening_call_username:
                            evening_call_username = username
                            has_evening_assignment = True
                        if username and 'N' in markers and not night_call_username:
                            night_call_username = username
                            has_night_assignment = True
                    
                    # Check if we have any D or 2C markers at all (even if unknown)
                    if not has_day_assignment:
                        for username, markers in day_data['staff']:
                            if 'D' in markers or '2C' in markers:
                                has_day_assignment = True
                                if username and '2C' in markers and not day_call_username:
                                    day_call_username = username
                                break
                    
                    # Check if we have any E markers at all (even if unknown)
                    if not has_evening_assignment:
                        for username, markers in day_data['staff']:
                            if 'E' in markers:
                                has_evening_assignment = True
                                break
                    
                    # Check if we have any N markers at all (even if unknown)
                    if not has_night_assignment:
                        for username, markers in day_data['staff']:
                            if 'N' in markers:
                                has_night_assignment = True
                                break
                    
                    # Track expected and create entries
                    # Day call
                    if has_day_assignment:
                        self.debug_tracker.add_expected_entry('Cardiology', day, '700', '1600', 'Team123', None)
                        if day_call_username:
                            role = self._get_employee_role(day_call_username)
                            entry = create_schedule_entry(day_call_username, '123', current_date, '700', next_date, '1600', role)
                            entry['NOTES'] = '1st Day Call'
                            output_rows.append(entry)
                            self.debug_tracker.mark_entry_generated('Cardiology', day, '700', '1600')
                    
                    # Evening call
                    if has_evening_assignment:
                        self.debug_tracker.add_expected_entry('Cardiology', day, '1600', '1900', 'Team123', None)
                        if evening_call_username:
                            role = self._get_employee_role(evening_call_username)
                            entry = create_schedule_entry(evening_call_username, '123', current_date, '1600', next_date, '1900', role)
                            entry['NOTES'] = 'Evening Call'
                            output_rows.append(entry)
                            self.debug_tracker.mark_entry_generated('Cardiology', day, '1600', '1900')
                    
                    # Night call
                    if has_night_assignment:
                        self.debug_tracker.add_expected_entry('Cardiology', day, '1900', '700', 'Team123', None)
                        if night_call_username:
                            role = self._get_employee_role(night_call_username)
                            entry = create_schedule_entry(night_call_username, '123', current_date, '1900', next_date, '700', role)
                            entry['NOTES'] = 'Night Call'
                            output_rows.append(entry)
                            self.debug_tracker.mark_entry_generated('Cardiology', day, '1900', '700')
                else:
                    # WEEKENDS: 3 entries
                    # 1. Consultant 2nd Weekend Day Call (7:00-19:00)
                    # 2. Staff/Fellow 1st Weekend Day Call (7:00-19:00)
                    # 3. Staff/Fellow Night Call (19:00-07:00)
                    
                    # Find consultant with 'X' marker for weekends (skip None)
                    consultant_username = None
                    for username, markers in day_data['consultants']:
                        if username and 'X' in markers:
                            consultant_username = username
                            break
                    
                    if not consultant_username:
                        for username, markers in day_data['consultants']:
                            if username and 'D' in markers:
                                consultant_username = username
                                break
                    
                    if not consultant_username:
                        for username, markers in day_data['consultants']:
                            if username:
                                consultant_username = username
                                break
                    
                    # Find staff for Weekend (skip None)
                    day_call_username = None
                    night_call_username = None
                    has_day_assignment = False
                    has_night_assignment = False
                    
                    for username, markers in day_data['staff']:
                        if username and 'D' in markers and not day_call_username:
                            day_call_username = username
                            has_day_assignment = True
                        if username and 'N' in markers and not night_call_username:
                            night_call_username = username
                            has_night_assignment = True
                    
                    # Check for 2C if no D found, and check if ANY day assignment exists
                    if not has_day_assignment:
                        for username, markers in day_data['staff']:
                            if 'D' in markers or '2C' in markers:
                                has_day_assignment = True
                                if username and '2C' in markers and not day_call_username:
                                    day_call_username = username
                                break
                    
                    # Check if any night assignment exists (even if unknown)
                    if not has_night_assignment:
                        for username, markers in day_data['staff']:
                            if 'N' in markers:
                                has_night_assignment = True
                                break
                    
                    # Track expected and create entries
                    # Consultant
                    if day_data['consultants']:
                        self.debug_tracker.add_expected_entry('Cardiology', day, '700', '1900', 'Team123', None)
                        if consultant_username:
                            entry = create_schedule_entry(consultant_username, '123', current_date, '700', next_date, '1900', '3042457')
                            entry['NOTES'] = '2nd Weekend Day Call'
                            output_rows.append(entry)
                            self.debug_tracker.mark_entry_generated('Cardiology', day, '700', '1900')
                    
                    # Staff day call
                    if has_day_assignment:
                        self.debug_tracker.add_expected_entry('Cardiology', day, '700', '1900', 'Team123', None)
                        if day_call_username:
                            role = self._get_employee_role(day_call_username)
                            entry = create_schedule_entry(day_call_username, '123', current_date, '700', next_date, '1900', role)
                            entry['NOTES'] = '1st Weekend Day Call'
                            output_rows.append(entry)
                            self.debug_tracker.mark_entry_generated('Cardiology', day, '700', '1900')
                    
                    # Staff night call
                    if has_night_assignment:
                        self.debug_tracker.add_expected_entry('Cardiology', day, '1900', '700', 'Team123', None)
                        if night_call_username:
                            role = self._get_employee_role(night_call_username)
                            entry = create_schedule_entry(night_call_username, '123', current_date, '1900', next_date, '700', role)
                            entry['NOTES'] = 'Night Call'
                            output_rows.append(entry)
                            self.debug_tracker.mark_entry_generated('Cardiology', day, '1900', '700')
            
            # Process Interventional Cardiologist team (94) - SECOND in order
            if day in interventional_data:
                username = interventional_data[day]
                
                # Determine start time based on day of week
                start_time = '1600' if is_weekday_flag else '700'
                
                # Track expected entry (schedule structure only, not employee)
                self.debug_tracker.add_expected_entry('Interventional Cardiologist', day, start_time, '700', 'Team94', None)
                
                # Only create entry if username is valid
                if username:
                    role = self._get_employee_role(username)
                    entry = create_schedule_entry(username, '94', current_date, start_time, next_date, '700', role)
                    entry['NOTES'] = 'On Call'
                    output_rows.append(entry)
                    self.debug_tracker.mark_entry_generated('Interventional Cardiologist', day, start_time, '700')
            
            # Process Cardiovascular team (8) - THIRD in order
            if day in cardiovascular_data:
                for username, roles in cardiovascular_data[day]:
                    for role in roles:
                        # Track expected entry (schedule structure only, not employee)
                        self.debug_tracker.add_expected_entry('Cardiovascular', day, '700', '700', 'Team8', None)
                        
                        # Only create entry if username is valid
                        if username:
                            entry = create_schedule_entry(username, '8', current_date, '700', next_date, '700', role)
                            output_rows.append(entry)
                            self.debug_tracker.mark_entry_generated('Cardiovascular', day, '700', '700')
        
        return output_rows
    
    def _find_username_by_identifier(self, identifier):
        """Find username by initials or name (flexible matching)"""
        if not identifier:
            return None
        
        identifier = str(identifier).strip()
        
        # Create reverse lookups from the department's employee map
        INITIALS_TO_USERNAME = {v['emp_initials']: k for k, v in self.employee_map.items()}
        NAME_TO_USERNAME = {v['emp_name']: k for k, v in self.employee_map.items()}
        
        # Try exact match with initials (case-insensitive)
        identifier_upper = identifier.upper()
        if identifier_upper in INITIALS_TO_USERNAME:
            return INITIALS_TO_USERNAME[identifier_upper]
        
        # Try exact match with name
        if identifier in NAME_TO_USERNAME:
            return NAME_TO_USERNAME[identifier]
        
        # Try normalized name matching (without periods, case-insensitive)
        identifier_normalized = identifier.replace('.', '').strip()
        for name, username in NAME_TO_USERNAME.items():
            name_normalized = name.replace('.', '').strip()
            if identifier_normalized.lower() == name_normalized.lower():
                return username
        
        # Not found - track as unknown
        # Determine if it looks like initials (short, all caps) or a name
        if len(identifier) <= 4 and identifier.isupper():
            self.debug_tracker.add_unknown_initials(identifier)
        else:
            self.debug_tracker.add_unknown_name(identifier)
        
        return None
    
    def _get_employee_role(self, username):
        """Get the primary role for an employee"""
        if username in self.employee_map:
            return self.employee_map[username]['emp_roles'][0]
        return '72'  # Default role
