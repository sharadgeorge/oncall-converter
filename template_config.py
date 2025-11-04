"""
[Department Name] Department Configuration
==========================================
Configuration for converting [Department] Excel schedules to import format.

This file contains all [Department]-specific logic for:
- File validation
- Data extraction
- Schedule processing

To use: Place in department_configs/ folder or same directory as main script.

Note: Utilities (DepartmentConfig, col_letter_to_index, etc.) are injected
by the main script during dynamic loading. No need to import them.
"""

import openpyxl
import calendar
import re
from datetime import datetime, timedelta
from pathlib import Path

# Utilities are injected by main script:
# - DepartmentConfig (base class)
# - col_letter_to_index (convert 'A' -> 1, 'B' -> 2, etc.)
# - is_weekday (check if date is Sun-Thu)
# - extract_month_year_from_filename (parse month/year from filename)
# - create_schedule_entry (create standardized entry dict)
# - DebugTracker (track expected/generated entries and unknown names)

class TemplateDepartmentConfig(DepartmentConfig):
    """Configuration for [Department] department schedules"""
    
    def get_department_name(self):
        """Return department name - must match EMPLOYEE_MAP 'department' field"""
        return "Template"  # TODO: Change this
    
    def get_file_requirements(self):
        """Return list of required file descriptions for user prompts"""
        return [
            "Schedule File 1 (.xlsx)",  # TODO: Describe your files
            "Schedule File 2 (.xlsx)"   # Add or remove as needed
        ]
    
    def get_team_abbr_map(self):
        """Return team name abbreviations for display in warnings/reports"""
        return {
            'Team1': 'T1',  # TODO: Define your team abbreviations
            'Team2': 'T2'
        }
    
    def validate_and_configure(self, files, month, year):
        """
        Validate files and configure department settings
        
        Args:
            files: List of Path objects for Excel files
            month: Initial month (may be None)
            year: Initial year (may be None)
            
        Returns:
            tuple: (workbooks_dict, config_dict)
        """
        # TODO: Validate file count
        if len(files) != 2:  # Adjust based on your requirements
            raise ValueError("Template department requires exactly 2 files")
        
        file1, file2 = files[0], files[1]
        
        print("\n" + "=" * 60)
        print("Validating [Department] Schedule Files")
        print("=" * 60)
        
        # ================================================================
        # STEP 1: Load and validate Excel files
        # ================================================================
        
        # TODO: Load your Excel files
        print(f"\n  Loading File 1: {file1.name}")
        wb1 = openpyxl.load_workbook(file1, data_only=True)
        
        # TODO: Find or validate the correct worksheet
        if 'ExpectedSheetName' in wb1.sheetnames:
            ws1 = wb1['ExpectedSheetName']
            print(f"  ✓ Found 'ExpectedSheetName' sheet")
        else:
            # Prompt user to select sheet
            print(f"  Available sheets: {wb1.sheetnames}")
            sheet_name = input("  Enter the sheet name: ").strip()
            if sheet_name in wb1.sheetnames:
                ws1 = wb1[sheet_name]
                print(f"  ✓ Using sheet '{sheet_name}'")
            else:
                raise ValueError(f"Sheet '{sheet_name}' not found")
        
        # TODO: Repeat for other files
        print(f"\n  Loading File 2: {file2.name}")
        wb2 = openpyxl.load_workbook(file2, data_only=True)
        ws2 = wb2.active  # Or use specific sheet name
        print(f"  ✓ Loaded File 2")
        
        # ================================================================
        # STEP 2: Parse and confirm month/year
        # ================================================================
        
        # Try to detect month/year from filename
        detected_month, detected_year = extract_month_year_from_filename(file1.stem)
        
        if detected_month:
            print(f"\n  Detected month: {calendar.month_name[detected_month]} {detected_year}")
            confirm = input(f"  Is this correct? (Y/n): ").strip().lower()
            if confirm in ['n', 'no']:
                month_input = input("  Enter month (name or number) and year: ").strip()
                month = None
                year = detected_year  # Keep detected year as fallback
                try:
                    # Check both full names and abbreviations
                    for m in range(1, 13):
                        month_name_lower = calendar.month_name[m].lower()
                        month_abbr_lower = calendar.month_abbr[m].lower()
                        if month_name_lower in month_input.lower() or month_abbr_lower in month_input.lower():
                            month = m
                            break
                    # Extract year
                    year_match = re.search(r'20\d{2}', month_input)
                    if year_match:
                        year = int(year_match.group())
                    # Fallback if month not found
                    if month is None:
                        print("  ⚠ Could not parse month, using detected values")
                        month = detected_month
                except:
                    print("  ⚠ Could not parse input, using detected values")
                    month = detected_month
            else:
                month = detected_month
                year = detected_year
        else:
            # No detection - prompt user
            month_input = input("  Enter month (name or number) and year: ").strip()
            month = None
            year = datetime.now().year
            try:
                for m in range(1, 13):
                    month_name_lower = calendar.month_name[m].lower()
                    month_abbr_lower = calendar.month_abbr[m].lower()
                    if month_name_lower in month_input.lower() or month_abbr_lower in month_input.lower():
                        month = m
                        break
                year_match = re.search(r'20\d{2}', month_input)
                if year_match:
                    year = int(year_match.group())
                if month is None:
                    print("  ⚠ Could not parse month, using current month")
                    month = datetime.now().month
            except:
                month = datetime.now().month
                year = datetime.now().year
        
        print(f"\n  ✓ Processing: {calendar.month_name[month]} {year}")
        
        # ================================================================
        # STEP 3: Configure row/column settings (optional)
        # ================================================================
        
        # TODO: If your schedules have configurable rows/columns, prompt here
        print("\nConfiguration:")
        print("-" * 60)
        
        # Example: prompt for data row range
        default_start_row = 5
        default_end_row = 20
        user_input = input(f"Data row range [default: {default_start_row}-{default_end_row}]: ").strip()
        
        if user_input and '-' in user_input:
            try:
                start, end = user_input.split('-')
                data_rows = (int(start), int(end))
                print(f"✓ Using rows {start}-{end}")
            except:
                data_rows = (default_start_row, default_end_row)
                print(f"✓ Using default rows")
        else:
            data_rows = (default_start_row, default_end_row)
            print(f"✓ Using default rows {default_start_row}-{default_end_row}")
        
        print()
        
        # ================================================================
        # STEP 4: Define team configurations
        # ================================================================
        
        # TODO: Define your teams and their settings
        teams = {
            'Team1': {
                'team_id': '100',  # System team ID
                'data_rows': data_rows,
                'columns': ['A', 'B', 'C'],
                # Add any other team-specific config
            },
            'Team2': {
                'team_id': '200',
                'data_rows': (10, 15),
                'columns': ['D', 'E'],
            }
        }
        
        # ================================================================
        # STEP 5: Package workbooks and config
        # ================================================================
        
        workbooks = {
            'file1': wb1,
            'file2': wb2,
            'ws1': ws1,  # You can include worksheets too
            'ws2': ws2
        }
        
        config = {
            'teams': teams,
            'month': month,
            'year': year,
            # Add any other configuration data
        }
        
        return workbooks, config
    
    def extract_schedule_data(self, workbooks, config, month, year):
        """
        Extract schedule data from workbooks
        
        Args:
            workbooks: Dict containing workbook/worksheet objects
            config: Dict containing configuration
            month: Month number (1-12)
            year: Year (e.g., 2025)
            
        Returns:
            list: List of schedule entry dictionaries
        """
        teams = config['teams']
        output_data = []
        
        # TODO: Implement your extraction logic
        
        print("\nExtracting schedule data...")
        
        # Loop through each team
        for team_name, team_config in teams.items():
            print(f"  Processing {team_name}...")
            
            # Extract data for this team
            team_data = self._extract_team_data(workbooks, team_config, month, year)
            
            # Convert to output entries
            team_entries = self._create_team_entries(team_name, team_data, year, month)
            
            output_data.extend(team_entries)
            print(f"    ✓ Generated {len(team_entries)} entries")
        
        print(f"\n✓ Total entries generated: {len(output_data)}")
        
        # Show sample
        if output_data:
            print("\nSample of first 3 entries:")
            for i, entry in enumerate(output_data[:3]):
                print(f"  {i+1}. {entry['EMPLOYEE']} | Team {entry['TEAM']} | {entry['STARTDATE']} {entry['STARTTIME']}-{entry['ENDTIME']}")
        
        return output_data
    
    def _extract_team_data(self, workbooks, team_config, month, year):
        """
        Extract raw data for a specific team
        
        TODO: Implement your data extraction logic
        
        Returns:
            dict: {day: [(username, metadata), ...]}
        """
        # Example structure - adjust for your needs
        team_data = {}
        days_in_month = calendar.monthrange(year, month)[1]
        
        # TODO: Read from Excel and populate team_data
        # Example:
        # for day in range(1, days_in_month + 1):
        #     team_data[day] = []
        #     
        #     # Find assignments for this day
        #     username = self._find_employee_for_day(workbooks, day, team_config)
        #     
        #     if username:
        #         team_data[day].append((username, {}))
        #         
        #         # Track expected entry
        #         self.debug_tracker.add_expected_entry(
        #             team_name='Team1',
        #             day=day,
        #             start_time='700',
        #             end_time='1600',
        #             source='Excel',
        #             employee=username
        #         )
        
        return team_data
    
    def _create_team_entries(self, team_name, team_data, year, month):
        """
        Convert extracted team data to standardized schedule entries
        
        Args:
            team_name: Name of the team
            team_data: Extracted data {day: [(username, metadata), ...]}
            year: Year
            month: Month
            
        Returns:
            list: List of entry dictionaries
        """
        entries = []
        team_id = '100'  # TODO: Get from team config
        
        for day, assignments in team_data.items():
            current_date = datetime(year, month, day)
            next_date = current_date + timedelta(days=1)
            
            for username, metadata in assignments:
                # TODO: Determine times based on your logic
                start_time = '700'
                end_time = '1600'
                
                # Get employee role
                role = self._get_employee_role(username)
                
                # Create entry
                entry = create_schedule_entry(
                    username,
                    team_id,
                    current_date,
                    start_time,
                    next_date,
                    end_time,
                    role
                )
                
                # Add notes if needed
                entry['NOTES'] = 'Day Shift'
                
                entries.append(entry)
                
                # Mark as generated
                self.debug_tracker.mark_entry_generated(
                    team_name,
                    day,
                    start_time,
                    end_time
                )
        
        return entries
    
    def _find_username_by_identifier(self, identifier):
        """
        Find username by initials or name (flexible matching)
        
        Args:
            identifier: Employee name or initials from Excel
            
        Returns:
            str: Username (employee ID) or None
        """
        if not identifier:
            return None
        
        identifier = str(identifier).strip()
        
        # Create reverse lookups from department's employee map
        INITIALS_TO_USERNAME = {v['emp_initials']: k for k, v in self.employee_map.items()}
        NAME_TO_USERNAME = {v['emp_name']: k for k, v in self.employee_map.items()}
        
        # Try exact match with initials (case-insensitive)
        identifier_upper = identifier.upper()
        if identifier_upper in INITIALS_TO_USERNAME:
            return INITIALS_TO_USERNAME[identifier_upper]
        
        # Try exact match with name
        if identifier in NAME_TO_USERNAME:
            return NAME_TO_USERNAME[identifier]
        
        # Try normalized name matching (without periods, case-insensitive)
        identifier_normalized = identifier.replace('.', '').strip()
        for name, username in NAME_TO_USERNAME.items():
            name_normalized = name.replace('.', '').strip()
            if identifier_normalized.lower() == name_normalized.lower():
                return username
        
        # Not found - track as unknown
        if len(identifier) <= 4 and identifier.isupper():
            self.debug_tracker.add_unknown_initials(identifier)
        else:
            self.debug_tracker.add_unknown_name(identifier)
        
        return None
    
    def _get_employee_role(self, username):
        """
        Get the primary role for an employee
        
        Args:
            username: Employee username/ID
            
        Returns:
            str: Role ID
        """
        if username in self.employee_map:
            return self.employee_map[username]['emp_roles'][0]
        return '1056'  # TODO: Set appropriate default role for your department
