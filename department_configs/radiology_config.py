"""
Radiology Department Configuration
===================================
Configuration for converting Radiology Excel schedules to import format.

This file contains all Radiology-specific logic for:
- File validation
- Data extraction
- Schedule processing

To use: Place in department_configs/ folder or same directory as main script.

Note: Utilities (DepartmentConfig, col_letter_to_index, etc.) are injected
by the main script during dynamic loading. No need to import them.
"""

import openpyxl
import calendar
import re
from datetime import datetime, timedelta
from pathlib import Path

# Utilities are injected by main script:
# - DepartmentConfig
# - col_letter_to_index
# - is_weekday
# - extract_month_year_from_filename
# - create_schedule_entry
# - DebugTracker

class RadiologyConfig(DepartmentConfig):
    """Configuration for Radiology department schedules"""
    
    def get_department_name(self):
        """Return department name"""
        return "Radiology"
    
    def get_file_requirements(self):
        """Return list of required file descriptions"""
        return [
            "Work Schedule file (.xlsx)",
            "OnCall Schedule file (.xlsx)"
        ]
    
    def get_team_abbr_map(self):
        """Return team name abbreviations for display"""
        return {
            'Gen_CT': 'GEN',
            'IRA': 'IRA',
            'MRI': 'MRI',
            'US': 'US',
            'Fluoro': 'FLU'
        }
    
    def get_default_row_config(self):
        """
        Return default row configuration for UI
        Radiology uses standard row layouts, no custom configuration needed
        """
        return None
    
    def get_default_sheet_names(self):
        """
        Return default sheet names for each file
        
        Returns:
            dict: {file_index: sheet_name}
        """
        return {
            0: 'WORK SCHEDULE',  # File 1: Work Schedule
            1: 'Sheet1'          # File 2: OnCall Schedule
        }
    
    def validate_and_configure(self, files, month, year, custom_config=None, selected_sheets=None):
        """
        Validate Radiology files and configure settings
        
        Args:
            files: List of Path objects [work_schedule, oncall_schedule]
            month: Initial month (may be None)
            year: Initial year (may be None)
            
        Returns:
            tuple: (workbooks_dict, config_dict)
        """
        if len(files) != 2:
            raise ValueError("Radiology requires exactly 2 files")
        
        work_file, oncall_file = files[0], files[1]
        
        print("\n" + "=" * 60)
        print("Validating Radiology Schedule Files")
        print("=" * 60)
        
        # Load Work Schedule
        print(f"\n  Loading Work Schedule: {work_file.name}")
        wb_work = openpyxl.load_workbook(work_file)
        
        # Use selected sheet if provided, otherwise try default
        if selected_sheets and 0 in selected_sheets:
            sheet_name = selected_sheets[0]
            if sheet_name in wb_work.sheetnames:
                ws_work = wb_work[sheet_name]
                print(f"  ✓ Using selected sheet: '{sheet_name}'")
            else:
                ws_work = wb_work[wb_work.sheetnames[0]]
                print(f"  ⚠ Selected sheet not found, using: '{wb_work.sheetnames[0]}'")
        elif 'WORK SCHEDULE' in wb_work.sheetnames:
            ws_work = wb_work['WORK SCHEDULE']
            print(f"  ✓ Found 'WORK SCHEDULE' sheet")
        else:
            # For Streamlit, use first sheet if WORK SCHEDULE not found
            ws_work = wb_work[wb_work.sheetnames[0]]
            print(f"  ✓ Using sheet '{wb_work.sheetnames[0]}'")
        
        # Load OnCall Schedule
        print(f"\n  Loading OnCall Schedule: {oncall_file.name}")
        wb_oncall = openpyxl.load_workbook(oncall_file, data_only=True)
        
        # Use selected sheet if provided, otherwise try default
        if selected_sheets and 1 in selected_sheets:
            sheet_name = selected_sheets[1]
            if sheet_name in wb_oncall.sheetnames:
                ws_oncall = wb_oncall[sheet_name]
                print(f"  ✓ Using selected sheet: '{sheet_name}'")
            else:
                ws_oncall = wb_oncall[wb_oncall.sheetnames[0]]
                print(f"  ⚠ Selected sheet not found, using: '{wb_oncall.sheetnames[0]}'")
        elif 'Sheet1' in wb_oncall.sheetnames:
            ws_oncall = wb_oncall['Sheet1']
            print(f"  ✓ Found 'Sheet1' sheet")
        else:
            # For Streamlit, use first sheet if Sheet1 not found
            ws_oncall = wb_oncall[wb_oncall.sheetnames[0]]
            print(f"  ✓ Using sheet '{wb_oncall.sheetnames[0]}'")
        
        # Use month/year from Streamlit parameters if provided
        if month is not None and year is not None:
            # Parameters already set by Streamlit - use them
            print(f"\n  ✓ Processing: {calendar.month_name[month]} {year}")
        else:
            # Fallback: Detect from filename (for command-line use)
            detected_month, detected_year = extract_month_year_from_filename(oncall_file.stem)
            if detected_month:
                print(f"\n  Detected month: {calendar.month_name[detected_month]} {detected_year}")
                month = detected_month
                year = detected_year
                print(f"  ✓ Processing: {calendar.month_name[month]} {year}")
            else:
                # Last fallback: use current date
                month = datetime.now().month
                year = datetime.now().year
                print(f"\n  Using current date: {calendar.month_name[month]} {year}")

        
        # Team configurations for Radiology
        teams = {
            'Gen_CT': {
                'team_id': '114',
                'work_cols': ['H', 'I'],
                'oncall_rows': (5, 21)
            },
            'IRA': {
                'team_id': '115',
                'work_cols': ['M'],
                'oncall_rows': (24, 27)
            },
            'MRI': {
                'team_id': '116',
                'work_cols': ['C'],
                'oncall_rows': (30, 38)
            },
            'US': {
                'team_id': '126',
                'work_cols': ['E'],
                'oncall_rows': (5, 21)
            },
            'Fluoro': {
                'team_id': '127',
                'work_cols': ['O'],
                'oncall_rows': (5, 21)
            }
        }
        
        workbooks = {
            'work': wb_work,
            'work_sheet': ws_work,
            'oncall': wb_oncall,
            'oncall_sheet': ws_oncall
        }
        
        config = {
            'teams': teams,
            'month': month,
            'year': year
        }
        
        return workbooks, config
    
    def _get_employee_from_work_schedule(self, ws_work, day_num, col_letter):
        """
        Get employee initials from work schedule for a specific day and column
        
        Args:
            ws_work: Worksheet object for work schedule
            day_num: Day of month (1-31)
            col_letter: Column letter (e.g., 'H', 'I')
            
        Returns:
            str: Employee initials or None
        """
        initials_to_empid = {v['emp_initials']: k for k, v in self.employee_map.items()}
        
        # Row ranges where schedule data appears
        row_ranges = [(5, 9), (13, 17), (21, 25), (29, 33), (37, 41)]
        col_idx = col_letter_to_index(col_letter)
        
        for row_range in row_ranges:
            for row in range(row_range[0], row_range[1] + 1):
                day_cell = ws_work.cell(row=row, column=1).value
                
                if day_cell:
                    # Extract day number from cell
                    if isinstance(day_cell, datetime):
                        cell_day = day_cell.day
                    else:
                        day_str = str(day_cell).strip()
                        if '-' in day_str:
                            day_str = day_str.split('-')[0]
                        try:
                            cell_day = int(day_str)
                        except ValueError:
                            continue
                    
                    # Found the correct day
                    if cell_day == day_num:
                        emp_cell = ws_work.cell(row=row, column=col_idx).value
                        if emp_cell:
                            emp_str = str(emp_cell).strip().upper()
                            
                            # Handle combined readers (e.g., "AS/TELE", "MF/MM")
                            if '/' in emp_str:
                                readers = [r.strip() for r in emp_str.split('/') if r.strip()]
                                
                                # Try to find first non-TELE reader
                                for reader in readers:
                                    if reader != 'TELE' and reader in initials_to_empid:
                                        return reader
                                    elif reader != 'TELE' and reader not in initials_to_empid:
                                        self.debug_tracker.add_unknown_initials(reader)
                                
                                # If all are TELE or invalid, return TELE if it exists
                                if 'TELE' in readers and 'TELE' in initials_to_empid:
                                    return 'TELE'
                            else:
                                # Single reader
                                if emp_str in initials_to_empid:
                                    return emp_str
                                else:
                                    self.debug_tracker.add_unknown_initials(emp_str)
        
        return None
    
    def _get_employee_from_oncall_schedule(self, ws_oncall, day_num, row_start, row_end):
        """
        Get employee marked with X from oncall schedule
        
        Args:
            ws_oncall: Worksheet object for oncall schedule
            day_num: Day of month (1-31)
            row_start: Starting row for this team
            row_end: Ending row for this team
            
        Returns:
            str: Employee initials or None
        """
        day_col = 3 + day_num  # Day columns start at D (column 4)
        
        for row in range(row_start, row_end + 1):
            # Skip header rows
            if row in [23, 29]:
                continue
            
            name_cell = ws_oncall.cell(row=row, column=1).value
            if not name_cell or str(name_cell).strip() == '':
                continue
            
            cell_value = ws_oncall.cell(row=row, column=day_col).value
            if cell_value and str(cell_value).strip().upper() == 'X':
                full_name = str(name_cell).strip().upper()
                original_name = str(name_cell).strip()
                
                # Try to match name to employee
                for emp_id, emp_data in self.employee_map.items():
                    emp_name_upper = emp_data['emp_name'].upper()
                    
                    # Handle LASTNAME, FIRSTNAME format
                    if ',' in full_name:
                        parts = full_name.split(',')
                        last_name = parts[0].strip()
                        first_name = parts[1].strip() if len(parts) > 1 else ''
                        
                        if last_name in emp_name_upper:
                            if first_name and first_name in emp_name_upper:
                                return emp_data['emp_initials']
                            elif not first_name:
                                return emp_data['emp_initials']
                    else:
                        # Standard format: "Dr. Firstname Lastname"
                        name_to_match = full_name.replace('DR.', '').replace('DR', '').strip()
                        emp_name_to_match = emp_name_upper.replace('DR.', '').replace('DR', '').strip()
                        
                        if name_to_match in emp_name_to_match or emp_name_to_match in name_to_match:
                            return emp_data['emp_initials']
                
                # No match found
                self.debug_tracker.add_unknown_name(original_name)
                return None
        
        return None
    
    def extract_schedule_data(self, workbooks, config, month, year):
        """
        Extract Radiology schedule data from workbooks
        
        Args:
            workbooks: Dict containing worksheet objects
            config: Dict containing configuration
            month: Month number (1-12)
            year: Year (e.g., 2025)
            
        Returns:
            list: List of schedule entry dictionaries
        """
        output_data = []
        ws_work = workbooks['work_sheet']
        ws_oncall = workbooks['oncall_sheet']
        teams = config['teams']
        
        # Create lookup: initials -> employee id
        initials_to_empid = {v['emp_initials']: k for k, v in self.employee_map.items()}
        
        days_in_month = calendar.monthrange(year, month)[1]
        
        print("\n" + "=" * 60)
        print(f"Extracting {self.get_department_name()} Schedule Data")
        print("=" * 60)
        
        # Process each day of the month
        for day in range(1, days_in_month + 1):
            current_date = datetime(year, month, day)
            next_date = current_date + timedelta(days=1)
            is_weekday_flag = is_weekday(current_date)
            
            # Process each team
            for team_name, team_config in teams.items():
                team_id = team_config['team_id']
                work_cols = team_config['work_cols']
                oncall_row_start, oncall_row_end = team_config['oncall_rows']
                
                if is_weekday_flag:
                    # Weekdays (Sun-Thu): Process work schedule + oncall
                    
                    if team_name == 'Gen_CT':
                        # Gen_CT has 3 blocks on weekdays: 0700-1100, 1100-1530, 1530-0700
                        
                        # Block 1: 0700-1100 (Column H)
                        emp1 = self._get_employee_from_work_schedule(ws_work, day, work_cols[0])
                        self.debug_tracker.add_expected_entry(team_name, day, '700', '1100', 'work', emp1)
                        if emp1 and emp1 in initials_to_empid:
                            entry = create_schedule_entry(
                                initials_to_empid[emp1], team_id, current_date, '700',
                                current_date, '1100', self.employee_map[initials_to_empid[emp1]]['emp_roles'][0]
                            )
                            output_data.append(entry)
                            self.debug_tracker.mark_entry_generated(team_name, day, '700', '1100')
                        
                        # Block 2: 1100-1530 (Column I)
                        emp2 = self._get_employee_from_work_schedule(ws_work, day, work_cols[1])
                        self.debug_tracker.add_expected_entry(team_name, day, '1100', '1530', 'work', emp2)
                        if emp2 and emp2 in initials_to_empid:
                            entry = create_schedule_entry(
                                initials_to_empid[emp2], team_id, current_date, '1100',
                                current_date, '1530', self.employee_map[initials_to_empid[emp2]]['emp_roles'][0]
                            )
                            output_data.append(entry)
                            self.debug_tracker.mark_entry_generated(team_name, day, '1100', '1530')
                        
                        # Block 3: 1530-0700 next day (OnCall)
                        emp3 = self._get_employee_from_oncall_schedule(ws_oncall, day, oncall_row_start, oncall_row_end)
                        self.debug_tracker.add_expected_entry(team_name, day, '1530', '700', 'oncall', emp3)
                        if emp3 and emp3 in initials_to_empid:
                            entry = create_schedule_entry(
                                initials_to_empid[emp3], team_id, current_date, '1530',
                                next_date, '700', self.employee_map[initials_to_empid[emp3]]['emp_roles'][0]
                            )
                            output_data.append(entry)
                            self.debug_tracker.mark_entry_generated(team_name, day, '1530', '700')
                    
                    else:
                        # Other teams have 2 blocks on weekdays: 0700-1530, 1530-0700
                        
                        # Block 1: 0700-1530 (Work schedule)
                        emp1 = self._get_employee_from_work_schedule(ws_work, day, work_cols[0])
                        self.debug_tracker.add_expected_entry(team_name, day, '700', '1530', 'work', emp1)
                        if emp1 and emp1 in initials_to_empid:
                            entry = create_schedule_entry(
                                initials_to_empid[emp1], team_id, current_date, '700',
                                current_date, '1530', self.employee_map[initials_to_empid[emp1]]['emp_roles'][0]
                            )
                            output_data.append(entry)
                            self.debug_tracker.mark_entry_generated(team_name, day, '700', '1530')
                        
                        # Block 2: 1530-0700 next day (OnCall)
                        emp2 = self._get_employee_from_oncall_schedule(ws_oncall, day, oncall_row_start, oncall_row_end)
                        self.debug_tracker.add_expected_entry(team_name, day, '1530', '700', 'oncall', emp2)
                        if emp2 and emp2 in initials_to_empid:
                            entry = create_schedule_entry(
                                initials_to_empid[emp2], team_id, current_date, '1530',
                                next_date, '700', self.employee_map[initials_to_empid[emp2]]['emp_roles'][0]
                            )
                            output_data.append(entry)
                            self.debug_tracker.mark_entry_generated(team_name, day, '1530', '700')
                
                else:
                    # Weekends (Fri-Sat): Only oncall, full 24 hours 0700-0700
                    emp = self._get_employee_from_oncall_schedule(ws_oncall, day, oncall_row_start, oncall_row_end)
                    self.debug_tracker.add_expected_entry(team_name, day, '700', '700', 'oncall', emp)
                    if emp and emp in initials_to_empid:
                        entry = create_schedule_entry(
                            initials_to_empid[emp], team_id, current_date, '700',
                            next_date, '700', self.employee_map[initials_to_empid[emp]]['emp_roles'][0]
                        )
                        output_data.append(entry)
                        self.debug_tracker.mark_entry_generated(team_name, day, '700', '700')
        
        return output_data
